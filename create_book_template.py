import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_excel_template(save_path: str) -> tuple[bool, str]:
    """
    Generate template Excel untuk import buku massal.
    Return: (success_boolean, error_message)
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template Import Buku"

        headers = [
            "Kode Buku",        # WAJIB — contoh: BK-2026-0001
            "Judul",            # WAJIB
            "Pengarang",
            "ISBN",
            "Kategori",
            "Penerbit",
            "Tahun Terbit",
            "Jumlah Eksemplar",
            "Lokasi Rak",
            "Keterangan",
        ]

        col_widths = [16, 40, 25, 18, 18, 25, 14, 18, 14, 30]

        # Style header
        header_font  = Font(bold=True, color="F9F6EE", size=11)
        header_fill  = PatternFill("solid", fgColor="4A5E3A")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border  = Border(
            bottom=Side(style="thin", color="C5B99A"),
            right=Side(style="thin", color="C5B99A"),
        )

        ws.row_dimensions[1].height = 30

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = width

        # Baris contoh
        sample_rows = [
            ["BK-2026-0001", "Laskar Pelangi",       "Andrea Hirata",   "9789799014054", "Fiksi",         "Bentang Pustaka", "2005", 3, "Rak A-1", ""],
            ["BK-2026-0002", "Bumi Manusia",          "Pramoedya A.T.",  "9789799024589", "Fiksi",         "Lentera Dipantara","1980", 2, "Rak A-1", "Tetralogi Buru"],
            ["BK-2026-0003", "Fisika Dasar Jilid 1",  "Halliday",        "9780471320579", "Sains",         "Erlangga",        "2010", 5, "Rak B-2", "Edisi ke-9"],
            ["BK-2026-0004", "Sejarah Indonesia",     "Sartono K.",      "",              "Sejarah",       "Gajah Mada UP",   "1998", 2, "Rak C-3", ""],
            ["BK-2026-0005", "Kamus Besar Bahasa Ind","Tim Redaksi",     "9789794079065", "Referensi",     "Balai Pustaka",   "2008", 1, "Rak D-1", "KBBI Edisi ke-4"],
        ]

        data_align = Alignment(vertical="center")
        data_fill_alt = PatternFill("solid", fgColor="E8E3D8")

        for row_idx, row_data in enumerate(sample_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_align
                if row_idx % 2 == 0:
                    cell.fill = data_fill_alt

        # Freeze baris header
        ws.freeze_panes = "A2"

        # Sheet petunjuk
        ws2 = wb.create_sheet("Petunjuk")
        ws2.column_dimensions["A"].width = 80

        petunjuk = [
            ("PETUNJUK PENGISIAN TEMPLATE IMPORT BUKU", True),
            ("", False),
            ("Kolom WAJIB (harus diisi):", True),
            ("  • Kode Buku  — Kode unik buku. Format bebas, disarankan: BK-YYYY-NNNN", False),
            ("                Contoh: BK-2026-0001", False),
            ("  • Judul      — Judul lengkap buku", False),
            ("", False),
            ("Kolom OPSIONAL (boleh dikosongkan):", True),
            ("  • Pengarang       — Nama pengarang / penulis / editor", False),
            ("  • ISBN            — Nomor ISBN (10 atau 13 digit)", False),
            ("  • Kategori        — Kategori buku (Fiksi, Sains, Sejarah, dll.)", False),
            ("  • Penerbit        — Nama penerbit", False),
            ("  • Tahun Terbit    — Tahun diterbitkan (4 digit)", False),
            ("  • Jumlah Eksemplar— Jumlah fisik buku yang dimiliki (default: 1)", False),
            ("  • Lokasi Rak      — Posisi buku di rak (contoh: Rak A-3)", False),
            ("  • Keterangan      — Catatan tambahan", False),
            ("", False),
            ("CATATAN PENTING:", True),
            ("  • Hapus baris contoh sebelum import (atau biarkan — jika kode duplikat akan dilewati)", False),
            ("  • Jangan mengubah nama kolom header (baris pertama)", False),
            ("  • Kode Buku yang sudah ada di database akan dilewati (tidak di-overwrite)", False),
            ("  • Gunakan menu '📥 Import Excel' di tab Buku untuk mulai import", False),
        ]

        ws2.cell(row=1, column=1).font = Font(bold=True, size=14, color="4A5E3A")
        for row_idx, (text, bold) in enumerate(petunjuk, start=1):
            cell = ws2.cell(row=row_idx, column=1, value=text)
            cell.font = Font(bold=bold, size=11 if bold else 10)
            cell.alignment = Alignment(vertical="center")

        wb.save(save_path)
        return True, ""

    except Exception as e:
        return False, str(e)

# (Opsional) Tetap bisa dijalankan langsung untuk testing
if __name__ == "__main__":
    ok, err = generate_excel_template("template_import_buku.xlsx")
    if ok:
        print("✔ Template berhasil dibuat: template_import_buku.xlsx")
    else:
        print(f"✘ Gagal: {err}")