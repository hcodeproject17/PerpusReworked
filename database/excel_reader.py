"""
database/excel_reader.py — Handler baca data anggota dari .xlsx
Menggunakan openpyxl langsung (tanpa pandas) agar build .exe lebih ringan.
"""

import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from config import (
    MEMBER_EXCEL_PATH,
    BACKUP_DIR,
    EXCEL_COL_BARCODE,
    EXCEL_COL_NAME,
)

logger = logging.getLogger(__name__)

# Cache sederhana: (path, mtime) → list[dict]
_cache: dict = {}


def _load_raw() -> list[dict]:
    """
    Baca seluruh baris dari anggota.xlsx.
    Hasil di-cache berdasarkan waktu modifikasi file agar tidak baca ulang
    setiap kali ada scan barcode.

    Returns:
        list of dict dengan key sesuai header kolom Excel
    Raises:
        FileNotFoundError : jika file Excel tidak ditemukan
        ValueError        : jika kolom wajib tidak ada
    """
    path = MEMBER_EXCEL_PATH

    if not path.exists():
        raise FileNotFoundError(f"File anggota tidak ditemukan: {path}")

    mtime = path.stat().st_mtime

    # Kembalikan cache jika file belum berubah
    if _cache.get("mtime") == mtime:
        return _cache["data"]

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    # Baca header dari baris pertama
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Validasi kolom wajib
    for required in (EXCEL_COL_BARCODE, EXCEL_COL_NAME):
        if required not in headers:
            wb.close()
            raise ValueError(
                f"Kolom '{required}' tidak ditemukan di {path.name}. "
                f"Kolom tersedia: {headers}"
            )

    # Baca semua baris data
    members: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))

        # Skip baris kosong
        barcode = record.get(EXCEL_COL_BARCODE)
        nama    = record.get(EXCEL_COL_NAME)
        if not barcode or not nama:
            continue

        record[EXCEL_COL_BARCODE] = str(barcode).strip()
        record[EXCEL_COL_NAME]    = str(nama).strip()
        members.append(record)

    wb.close()

    # Simpan ke cache
    _cache["mtime"] = mtime
    _cache["data"]  = members

    logger.info("Excel dimuat: %d anggota dari %s", len(members), path.name)
    return members


def load_members() -> list[dict]:
    """
    Kembalikan semua data anggota.

    Returns:
        list of dict — bisa kosong jika file kosong
    Raises:
        FileNotFoundError, ValueError (diteruskan dari _load_raw)
    """
    return _load_raw()

def read_source_excel(path: str | Path) -> list[dict]:
    """
    Baca file Excel sumber untuk generate kartu massal.
    Kolom wajib: 'Nama'
    """
    from pathlib import Path as _Path
    path = _Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File tidak ditemukan: {path}")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]

    if "Nama" not in headers:
        wb.close()
        raise ValueError(f"Kolom 'Nama' tidak ditemukan. Kolom tersedia: {headers}")

    members = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        nama = str(rec.get("Nama", "")).strip()
        if not nama or nama.lower() == "none":
            continue
        rec["Nama"] = nama
        members.append(rec)

    wb.close()
    return members

def find_by_barcode(barcode_id: str) -> Optional[dict]:
    """
    Cari anggota berdasarkan ID Barcode (exact match, case-insensitive).

    Returns:
        dict anggota jika ditemukan, None jika tidak
    """
    target = barcode_id.strip().upper()
    for member in _load_raw():
        if member[EXCEL_COL_BARCODE].upper() == target:
            return member
    return None


def find_by_name(name: str) -> list[dict]:
    """
    Cari anggota berdasarkan nama (partial match, case-insensitive).

    Returns:
        list of dict — bisa kosong
    """
    query = name.strip().lower()
    return [
        m for m in _load_raw()
        if query in m[EXCEL_COL_NAME].lower()
    ]


def search(query: str) -> list[dict]:
    """
    Cari anggota berdasarkan barcode ATAU nama (untuk search box GUI).
    Prioritas: exact barcode match dulu, lalu partial name match.

    Returns:
        list of dict — bisa kosong
    """
    query = query.strip()
    if not query:
        return []

    # Coba exact barcode dulu
    by_barcode = find_by_barcode(query)
    if by_barcode:
        return [by_barcode]

    # Fallback ke partial name
    return find_by_name(query)


def extra_member_columns(members: list[dict]) -> list[str]:
    """Kolom selain ID Barcode & Nama yang muncul di data anggota, urut
    kemunculan pertama (mis. "Kelas", "NISN"). Dipakai gui/member_tab.py
    (kolom tabel dinamis) dan gui/card_tab.py (filter per kolom)."""
    extra: list[str] = []
    for m in members:
        for k in m.keys():
            if k not in (EXCEL_COL_BARCODE, EXCEL_COL_NAME) and k not in extra:
                extra.append(k)
    return extra


# ══════════════════════════════════════════════════════════════════════════════
# CRUD tulis — dipakai oleh tab Daftar Anggota (gui/member_tab.py)
#
# Catatan: menghapus/mengganti ID seorang anggota di sini TIDAK menghapus
# riwayat kunjungan (sqlite_db) atau peminjaman (loan_db) yang sudah tercatat
# atas namanya — riwayat lama menyimpan nama+ID sebagai teks lepas (snapshot
# saat transaksi terjadi), bukan foreign key ke anggota.xlsx.
# ══════════════════════════════════════════════════════════════════════════════

def generate_next_member_id(year: Optional[int] = None) -> str:
    """
    ID Barcode berikutnya format "{YYYY}-{NNNN}", melanjutkan nomor terbesar
    di tahun yang sama. Dipakai tombol "Auto" di form tambah anggota.
    """
    year = year or datetime.now().year
    prefix = f"{year}-"
    try:
        existing_ids = [m[EXCEL_COL_BARCODE] for m in _load_raw()]
    except FileNotFoundError:
        existing_ids = []

    year_nums = [
        int(bid[len(prefix):])
        for bid in existing_ids
        if bid.startswith(prefix) and bid[len(prefix):].isdigit()
    ]
    next_num = (max(year_nums) + 1) if year_nums else 1
    return f"{year}-{next_num:04d}"


def _assign_member_ids(
    new_members: list[dict],
    year: Optional[int] = None,
    id_column: Optional[str] = None,
) -> tuple[list[dict], list[str]]:
    """
    Tentukan ID Barcode untuk tiap anggota baru hasil read_source_excel().

    Args:
        id_column : None = generate otomatis (YYYY-NNNN) untuk semua baris;
                    nama kolom = pakai nilai kolom itu sebagai ID Barcode
                    (mis. "NISN"). Baris yang nilainya kosong di kolom itu
                    tetap diberi ID otomatis (lihat warnings).

    Returns:
        (list anggota + key EXCEL_COL_BARCODE terisi, list pesan warning)
    """
    warnings: list[str] = []
    year = year or datetime.now().year
    try:
        existing_ids = {m[EXCEL_COL_BARCODE].upper() for m in _load_raw()}
    except FileNotFoundError:
        existing_ids = set()

    prefix = f"{year}-"
    next_auto = 1
    year_nums = [
        int(bid[len(prefix):]) for bid in existing_ids
        if bid.startswith(prefix) and bid[len(prefix):].isdigit()
    ]
    if year_nums:
        next_auto = max(year_nums) + 1

    result: list[dict] = []
    for member in new_members:
        nama = member.get(EXCEL_COL_NAME, "")
        barcode_id = None

        if id_column:
            raw = member.get(id_column)
            candidate = str(raw).strip() if raw is not None else ""
            if candidate:
                barcode_id = candidate
            else:
                warnings.append(f"'{nama}': kolom '{id_column}' kosong, ID di-generate otomatis.")

        if not barcode_id:
            while f"{year}-{next_auto:04d}".upper() in existing_ids:
                next_auto += 1
            barcode_id = f"{year}-{next_auto:04d}"
            next_auto += 1

        if barcode_id.upper() in existing_ids:
            warnings.append(f"'{nama}': ID '{barcode_id}' sudah dipakai, baris dilewati.")
            continue

        existing_ids.add(barcode_id.upper())
        member[EXCEL_COL_BARCODE] = barcode_id
        result.append(member)

    return result, warnings


def _save_members_bulk(members_with_id: list[dict]) -> int:
    """Tambahkan banyak anggota sekaligus ke anggota.xlsx, satu save di akhir
    (lebih cepat dari memanggil add_member() berulang saat impor massal)."""
    if not members_with_id:
        return 0

    wb, ws, headers = _open_or_create_workbook()

    # Kolom baru dari data import (mis. "Kelas", "NISN") ditambahkan sebagai
    # header baru kalau belum ada.
    for member in members_with_id:
        for key in member:
            if key not in headers:
                headers.append(key)
                cell = ws.cell(row=1, column=len(headers), value=key)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.alignment = Alignment(horizontal="center")

    for member in members_with_id:
        ws.append([member.get(h, "") for h in headers])

    wb.save(str(MEMBER_EXCEL_PATH))
    wb.close()
    invalidate_cache()
    return len(members_with_id)


def import_members_from_excel(
    path: str | Path,
    id_column: Optional[str] = None,
    year: Optional[int] = None,
) -> tuple[int, int, list[str]]:
    """
    Import massal anggota baru dari file Excel sumber (kolom wajib: "Nama").

    Args:
        id_column : None = ID Barcode digenerate otomatis; atau nama kolom
                    sumber yang dipakai sebagai ID (mis. "NISN").

    Returns:
        (jumlah berhasil, jumlah dilewati, list pesan warning/error)
    """
    try:
        new_members = read_source_excel(path)
    except (FileNotFoundError, ValueError) as exc:
        return 0, 0, [str(exc)]

    if not new_members:
        return 0, 0, ["Tidak ada data anggota (kolom 'Nama') di file sumber."]

    members_with_id, warnings = _assign_member_ids(new_members, year=year, id_column=id_column)
    ok = _save_members_bulk(members_with_id)
    err = len(new_members) - ok
    logger.info("Import anggota: %d berhasil, %d dilewati.", ok, err)
    return ok, err, warnings


def _open_or_create_workbook() -> tuple:
    """Buka anggota.xlsx, atau buat baru dengan header minimal jika belum ada.
    Returns (workbook, worksheet, headers)."""
    if MEMBER_EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(str(MEMBER_EXCEL_PATH))
        ws = wb.active
        headers = [
            str(c.value).strip() if c.value else ""
            for c in next(ws.iter_rows(max_row=1))
        ]
        if not headers or headers == [""]:
            headers = [EXCEL_COL_BARCODE, EXCEL_COL_NAME]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Anggota"
        headers = [EXCEL_COL_BARCODE, EXCEL_COL_NAME]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
    return wb, ws, headers


def add_member(
    nama: str,
    extra: Optional[dict] = None,
    barcode_id: Optional[str] = None,
) -> tuple[bool, str, Optional[dict]]:
    """
    Tambah satu anggota baru ke anggota.xlsx.

    Args:
        nama       : nama anggota (wajib)
        extra      : kolom tambahan opsional, mis. {"Kelas": "7A"}
        barcode_id : ID Barcode manual; None = generate otomatis (YYYY-NNNN)

    Returns:
        (ok, pesan, dict_anggota_baru_atau_None)
    """
    nama = nama.strip()
    if not nama:
        return False, "Nama tidak boleh kosong.", None

    wb, ws, headers = _open_or_create_workbook()

    col_barcode = headers.index(EXCEL_COL_BARCODE) + 1
    existing_ids = {
        str(row[col_barcode - 1].value).strip().upper()
        for row in ws.iter_rows(min_row=2)
        if row[col_barcode - 1].value
    }

    if barcode_id:
        barcode_id = barcode_id.strip()
        if not barcode_id:
            wb.close()
            return False, "ID Barcode tidak boleh kosong.", None
        if barcode_id.upper() in existing_ids:
            wb.close()
            return False, f"ID Barcode '{barcode_id}' sudah dipakai.", None
    else:
        barcode_id = generate_next_member_id()

    extra = extra or {}
    for key in extra:
        if key not in headers:
            headers.append(key)
            ws.cell(row=1, column=len(headers), value=key)

    row_data = {**extra, EXCEL_COL_BARCODE: barcode_id, EXCEL_COL_NAME: nama}
    ws.append([row_data.get(h, "") for h in headers])

    wb.save(str(MEMBER_EXCEL_PATH))
    wb.close()
    invalidate_cache()
    logger.info("Anggota baru ditambahkan: %s (%s)", nama, barcode_id)
    return True, "OK", dict(row_data)


def update_member(original_barcode_id: str, updates: dict) -> tuple[bool, str]:
    """
    Perbarui data anggota yang dicari lewat ID Barcode lamanya.
    `updates` boleh berisi EXCEL_COL_NAME, EXCEL_COL_BARCODE (untuk ganti ID),
    dan kolom lain apa pun (ditambahkan sebagai header baru kalau belum ada).
    """
    if not MEMBER_EXCEL_PATH.exists():
        return False, "File anggota.xlsx belum ada."

    wb = openpyxl.load_workbook(str(MEMBER_EXCEL_PATH))
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
    if EXCEL_COL_BARCODE not in headers:
        wb.close()
        return False, "Kolom ID Barcode tidak ditemukan di anggota.xlsx."
    col_barcode = headers.index(EXCEL_COL_BARCODE) + 1

    target = original_barcode_id.strip().upper()
    found_row = None
    for row in ws.iter_rows(min_row=2):
        val = row[col_barcode - 1].value
        if val and str(val).strip().upper() == target:
            found_row = row
            break

    if found_row is None:
        wb.close()
        return False, f"Anggota dengan ID '{original_barcode_id}' tidak ditemukan."

    new_barcode = updates.get(EXCEL_COL_BARCODE)
    if new_barcode is not None:
        new_barcode = str(new_barcode).strip()
        if not new_barcode:
            wb.close()
            return False, "ID Barcode tidak boleh dikosongkan."
        if new_barcode.upper() != target:
            for row in ws.iter_rows(min_row=2):
                val = row[col_barcode - 1].value
                if val and str(val).strip().upper() == new_barcode.upper():
                    wb.close()
                    return False, f"ID Barcode '{new_barcode}' sudah dipakai anggota lain."

    for key in updates:
        if key not in headers:
            headers.append(key)
            ws.cell(row=1, column=len(headers), value=key)

    for key, value in updates.items():
        col_idx = headers.index(key) + 1
        found_row[col_idx - 1].value = value

    wb.save(str(MEMBER_EXCEL_PATH))
    wb.close()
    invalidate_cache()
    logger.info("Anggota '%s' diperbarui.", original_barcode_id)
    return True, "OK"


def delete_member(barcode_id: str) -> tuple[bool, str]:
    """Hapus satu anggota berdasarkan ID Barcode (exact match)."""
    if not MEMBER_EXCEL_PATH.exists():
        return False, "File anggota.xlsx belum ada."

    wb = openpyxl.load_workbook(str(MEMBER_EXCEL_PATH))
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
    if EXCEL_COL_BARCODE not in headers:
        wb.close()
        return False, "Kolom ID Barcode tidak ditemukan di anggota.xlsx."
    col_barcode = headers.index(EXCEL_COL_BARCODE) + 1

    target = barcode_id.strip().upper()
    row_num = None
    for row in ws.iter_rows(min_row=2):
        val = row[col_barcode - 1].value
        if val and str(val).strip().upper() == target:
            row_num = row[0].row
            break

    if row_num is None:
        wb.close()
        return False, f"Anggota dengan ID '{barcode_id}' tidak ditemukan."

    ws.delete_rows(row_num, 1)
    wb.save(str(MEMBER_EXCEL_PATH))
    wb.close()
    invalidate_cache()
    logger.info("Anggota '%s' dihapus.", barcode_id)
    return True, "OK"


def delete_members_bulk(barcode_ids: list[str]) -> tuple[bool, str]:
    """Hapus beberapa anggota sekaligus (satu save di akhir, lebih cepat dari
    memanggil delete_member() berulang saat menghapus banyak baris)."""
    if not MEMBER_EXCEL_PATH.exists():
        return False, "File anggota.xlsx belum ada."
    if not barcode_ids:
        return True, "OK"

    wb = openpyxl.load_workbook(str(MEMBER_EXCEL_PATH))
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
    if EXCEL_COL_BARCODE not in headers:
        wb.close()
        return False, "Kolom ID Barcode tidak ditemukan di anggota.xlsx."
    col_barcode = headers.index(EXCEL_COL_BARCODE) + 1

    targets = {b.strip().upper() for b in barcode_ids}
    rows_to_delete = [
        row[0].row
        for row in ws.iter_rows(min_row=2)
        if row[col_barcode - 1].value
        and str(row[col_barcode - 1].value).strip().upper() in targets
    ]

    if not rows_to_delete:
        wb.close()
        return False, "Tidak ada anggota yang cocok untuk dihapus."

    # Hapus dari bawah ke atas supaya nomor baris yang belum dihapus tidak bergeser
    for row_num in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_num, 1)

    wb.save(str(MEMBER_EXCEL_PATH))
    wb.close()
    invalidate_cache()
    logger.info("%d anggota dihapus massal.", len(rows_to_delete))
    return True, f"{len(rows_to_delete)} anggota dihapus."


def backup_excel() -> Optional[Path]:
    """
    Buat salinan anggota.xlsx ke folder backup dengan timestamp.
    Dipanggil saat startup aplikasi.

    Returns:
        Path file backup jika berhasil, None jika gagal
    """
    src = MEMBER_EXCEL_PATH
    if not src.exists():
        logger.warning("Backup dibatalkan: file sumber tidak ada.")
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = BACKUP_DIR / f"anggota_backup_{timestamp}.xlsx"

    try:
        shutil.copy2(str(src), str(dst))
        logger.info("Backup Excel berhasil: %s", dst.name)
        return dst
    except OSError as exc:
        logger.error("Gagal backup Excel: %s", exc)
        return None


def invalidate_cache() -> None:
    """
    Paksa reload Excel pada akses berikutnya.
    Berguna jika file Excel diperbarui saat aplikasi berjalan.
    """
    _cache.clear()
    logger.debug("Cache Excel dikosongkan.")