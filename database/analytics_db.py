"""
database/analytics_db.py — Query analitik untuk tab analisis

Semua query dipisah dari sqlite_db.py agar tidak mencampur
logika operasional (catat kunjungan) dengan logika pelaporan.
"""

import sqlite3
import logging
from datetime import date, datetime, timedelta
from typing import Optional

from config import MEMBER_EXCEL_PATH, EXCEL_COL_BARCODE, EXCEL_COL_NAME
from database.connection import get_connection
import openpyxl

logger = logging.getLogger(__name__)


def _conn() -> sqlite3.Connection:
    return get_connection(wal=False, foreign_keys=False)


# ══════════════════════════════════════════════════════════════════════════════
# Helper
# ══════════════════════════════════════════════════════════════════════════════

def _date_range(start: date, end: date) -> tuple[str, str]:
    return start.isoformat(), end.isoformat()


# ══════════════════════════════════════════════════════════════════════════════
# Ringkasan
# ══════════════════════════════════════════════════════════════════════════════

def get_summary(start: date, end: date) -> dict:
    """
    Ringkasan statistik dalam rentang tanggal.

    Returns dict:
        total_visits    : int
        unique_visitors : int
        avg_per_day     : float
        busiest_day     : str (nama hari Indonesia)
        busiest_count   : int
        peak_hour       : int
        peak_hour_count : int
        days_with_data  : int
    """
    s, e = _date_range(start, end)
    try:
        with _conn() as conn:
            # Total kunjungan
            total = conn.execute(
                "SELECT COUNT(*) FROM histori_kunjungan WHERE tanggal BETWEEN ? AND ?",
                (s, e)
            ).fetchone()[0]

            # Pengunjung unik
            unique = conn.execute(
                "SELECT COUNT(DISTINCT barcode_id) FROM histori_kunjungan WHERE tanggal BETWEEN ? AND ?",
                (s, e)
            ).fetchone()[0]

            # Hari dengan kunjungan terbanyak
            busiest = conn.execute("""
                SELECT tanggal, COUNT(*) as cnt
                FROM histori_kunjungan
                WHERE tanggal BETWEEN ? AND ?
                GROUP BY tanggal
                ORDER BY cnt DESC LIMIT 1
            """, (s, e)).fetchone()

            # Jam tersibuk
            peak = conn.execute("""
                SELECT CAST(strftime('%H', waktu_masuk) AS INTEGER) as jam, COUNT(*) as cnt
                FROM histori_kunjungan
                WHERE tanggal BETWEEN ? AND ?
                GROUP BY jam
                ORDER BY cnt DESC LIMIT 1
            """, (s, e)).fetchone()

            # Hari aktif (hari yang ada kunjungannya)
            days_active = conn.execute(
                "SELECT COUNT(DISTINCT tanggal) FROM histori_kunjungan WHERE tanggal BETWEEN ? AND ?",
                (s, e)
            ).fetchone()[0]

        _DAYS_ID = ["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu"]

        busiest_day   = ""
        busiest_count = 0
        if busiest:
            d = date.fromisoformat(busiest["tanggal"])
            busiest_day   = _DAYS_ID[d.weekday()]
            busiest_count = busiest["cnt"]

        return {
            "total_visits":    total,
            "unique_visitors": unique,
            "avg_per_day":     round(total / max(days_active, 1), 1),
            "busiest_day":     busiest_day,
            "busiest_count":   busiest_count,
            "peak_hour":       peak["jam"] if peak else 0,
            "peak_hour_count": peak["cnt"] if peak else 0,
            "days_with_data":  days_active,
        }
    except Exception as exc:
        logger.error("get_summary error: %s", exc)
        return {
            "total_visits": 0, "unique_visitors": 0, "avg_per_day": 0.0,
            "busiest_day": "-", "busiest_count": 0,
            "peak_hour": 0, "peak_hour_count": 0, "days_with_data": 0,
        }


# ══════════════════════════════════════════════════════════════════════════════
# Tren kunjungan
# ══════════════════════════════════════════════════════════════════════════════

def get_daily_trend(start: date, end: date) -> list[dict]:
    """
    Jumlah kunjungan per hari dalam rentang tanggal.
    Hari tanpa kunjungan tetap muncul dengan count=0.

    Returns list[dict]: tanggal (str), count (int), label (str DD/MM)
    """
    s, e = _date_range(start, end)
    try:
        with _conn() as conn:
            rows = conn.execute("""
                SELECT tanggal, COUNT(*) as cnt
                FROM histori_kunjungan
                WHERE tanggal BETWEEN ? AND ?
                GROUP BY tanggal
                ORDER BY tanggal
            """, (s, e)).fetchall()

        # Isi hari kosong
        data_map = {r["tanggal"]: r["cnt"] for r in rows}
        result   = []
        cur = start
        while cur <= end:
            iso = cur.isoformat()
            result.append({
                "tanggal": iso,
                "count":   data_map.get(iso, 0),
                "label":   cur.strftime("%d/%m"),
                "hari":    ["Sen","Sel","Rab","Kam","Jum","Sab","Min"][cur.weekday()],
            })
            cur += timedelta(days=1)
        return result
    except Exception as exc:
        logger.error("get_daily_trend error: %s", exc)
        return []


def get_weekly_trend(start: date, end: date) -> list[dict]:
    """Agregasi kunjungan per minggu."""
    daily = get_daily_trend(start, end)
    weeks: dict[str, int] = {}
    for d in daily:
        dt  = date.fromisoformat(d["tanggal"])
        key = f"Mg {dt.isocalendar()[1]}/{dt.year}"
        weeks[key] = weeks.get(key, 0) + d["count"]
    return [{"label": k, "count": v} for k, v in weeks.items()]


def get_day_of_week_distribution(start: date, end: date) -> list[dict]:
    """
    Rata-rata kunjungan per hari-dalam-minggu (Senin–Minggu).
    """
    s, e = _date_range(start, end)
    _DAYS = ["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu"]
    try:
        with _conn() as conn:
            rows = conn.execute("""
                SELECT
                    CAST(strftime('%w', tanggal) AS INTEGER) as dow,
                    COUNT(*) as total,
                    COUNT(DISTINCT tanggal) as days
                FROM histori_kunjungan
                WHERE tanggal BETWEEN ? AND ?
                GROUP BY dow
            """, (s, e)).fetchall()

        # strftime %w: 0=Minggu … 6=Sabtu → konversi ke Mon=0
        dow_map = {}
        for r in rows:
            # %w: Sun=0,Mon=1,...,Sat=6 → kita mau Mon=0
            idx = (r["dow"] - 1) % 7
            dow_map[idx] = round(r["total"] / max(r["days"], 1), 1)

        return [
            {"hari": _DAYS[i], "avg": dow_map.get(i, 0.0)}
            for i in range(7)
        ]
    except Exception as exc:
        logger.error("get_day_of_week_distribution error: %s", exc)
        return [{"hari": d, "avg": 0.0} for d in _DAYS]


# ══════════════════════════════════════════════════════════════════════════════
# Per jam
# ══════════════════════════════════════════════════════════════════════════════

def get_hourly_distribution(start: date, end: date) -> list[dict]:
    """
    Distribusi kunjungan per jam (00–23).
    Returns list 24 elemen, jam tanpa kunjungan count=0.
    """
    s, e = _date_range(start, end)
    try:
        with _conn() as conn:
            rows = conn.execute("""
                SELECT
                    CAST(strftime('%H', waktu_masuk) AS INTEGER) as jam,
                    COUNT(*) as cnt
                FROM histori_kunjungan
                WHERE tanggal BETWEEN ? AND ?
                GROUP BY jam
                ORDER BY jam
            """, (s, e)).fetchall()

        hour_map = {r["jam"]: r["cnt"] for r in rows}
        return [
            {"jam": h, "label": f"{h:02d}:00", "count": hour_map.get(h, 0)}
            for h in range(24)
        ]
    except Exception as exc:
        logger.error("get_hourly_distribution error: %s", exc)
        return [{"jam": h, "label": f"{h:02d}:00", "count": 0} for h in range(24)]


# ══════════════════════════════════════════════════════════════════════════════
# Top anggota
# ══════════════════════════════════════════════════════════════════════════════

def get_top_visitors(start: date, end: date, limit: int = 20) -> list[dict]:
    """
    Daftar anggota dengan kunjungan terbanyak dalam rentang tanggal.

    Returns list[dict]: rank, barcode_id, nama, count, pct (% dari total)
    """
    s, e = _date_range(start, end)
    try:
        with _conn() as conn:
            rows = conn.execute("""
                SELECT barcode_id, nama, COUNT(*) as cnt
                FROM histori_kunjungan
                WHERE tanggal BETWEEN ? AND ?
                GROUP BY barcode_id
                ORDER BY cnt DESC
                LIMIT ?
            """, (s, e, limit)).fetchall()

            total = conn.execute(
                "SELECT COUNT(*) FROM histori_kunjungan WHERE tanggal BETWEEN ? AND ?",
                (s, e)
            ).fetchone()[0]

        result = []
        for i, r in enumerate(rows):
            result.append({
                "rank":       i + 1,
                "barcode_id": r["barcode_id"],
                "nama":       r["nama"],
                "count":      r["cnt"],
                "pct":        round(r["cnt"] / max(total, 1) * 100, 1),
            })
        return result
    except Exception as exc:
        logger.error("get_top_visitors error: %s", exc)
        return []


def get_all_visits_detail(start: date, end: date) -> list[dict]:
    """Semua baris kunjungan untuk export Excel."""
    s, e = _date_range(start, end)
    try:
        with _conn() as conn:
            rows = conn.execute("""
                SELECT barcode_id, nama, tanggal, waktu_masuk
                FROM histori_kunjungan
                WHERE tanggal BETWEEN ? AND ?
                ORDER BY waktu_masuk DESC
            """, (s, e)).fetchall()
        return [dict(r) for r in rows]
    except Exception as exc:
        logger.error("get_all_visits_detail error: %s", exc)
        return []


# ══════════════════════════════════════════════════════════════════════════════
# Anggota tidak aktif
# ══════════════════════════════════════════════════════════════════════════════

def get_inactive_members(since: Optional[date] = None) -> list[dict]:
    """
    Anggota yang terdaftar di anggota.xlsx tapi belum pernah ada
    di histori_kunjungan sejak tanggal 'since' (default: semua waktu).

    Returns list[dict]: barcode_id, nama, kolom_lain...
    """
    if not MEMBER_EXCEL_PATH.exists():
        return []
    try:
        # Baca semua anggota dari Excel
        wb = openpyxl.load_workbook(str(MEMBER_EXCEL_PATH), read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
        all_members = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = dict(zip(headers, row))
            bid  = str(rec.get(EXCEL_COL_BARCODE, "") or "").strip()
            nama = str(rec.get(EXCEL_COL_NAME, "")    or "").strip()
            if bid and nama:
                all_members.append(rec)
        wb.close()

        # Ambil semua barcode yang pernah masuk dari SQLite
        with _conn() as conn:
            if since:
                rows = conn.execute(
                    "SELECT DISTINCT barcode_id FROM histori_kunjungan WHERE tanggal >= ?",
                    (since.isoformat(),)
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT DISTINCT barcode_id FROM histori_kunjungan"
                ).fetchall()

        active_ids = {r["barcode_id"] for r in rows}

        return [
            m for m in all_members
            if m[EXCEL_COL_BARCODE] not in active_ids
        ]
    except Exception as exc:
        logger.error("get_inactive_members error: %s", exc)
        return []


# ══════════════════════════════════════════════════════════════════════════════
# Export Excel
# ══════════════════════════════════════════════════════════════════════════════

def export_to_excel(start: date, end: date, out_path: str) -> int:
    """
    Export laporan lengkap ke .xlsx dengan beberapa sheet.

    Sheet:
        1. Ringkasan       — stat card
        2. Detail kunjungan — semua baris
        3. Top anggota     — ranking
        4. Per hari        — trend harian
        5. Per jam         — distribusi jam
        6. Tidak aktif     — anggota belum pernah masuk

    Returns jumlah baris detail yang diekspor.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    HDR_FILL  = PatternFill("solid", fgColor="4A5E3A")
    HDR_FONT  = Font(bold=True, color="F9F6EE", size=11)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center")
    THIN      = Side(style="thin", color="C5B99A")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    TITLE_FONT = Font(bold=True, size=13, color="1E2818")
    SUB_FONT   = Font(size=10, color="6B7F58")

    def _hdr_row(ws, row_idx, cols):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=row_idx, column=ci, value=col)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = HDR_ALIGN; c.border = BORDER

    def _data_row(ws, row_idx, values, alt=False):
        alt_fill = PatternFill("solid", fgColor="E8E3D8") if alt else None
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.border = BORDER
            c.alignment = Alignment(vertical="center")
            if alt_fill:
                c.fill = alt_fill

    def _autowidth(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # ── Sheet 1: Ringkasan ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Ringkasan"
    summary = get_summary(start, end)

    ws1["A1"] = "LAPORAN PERPUSTAKAAN"
    ws1["A1"].font = TITLE_FONT
    ws1["A2"] = f"Periode: {start.strftime('%d %B %Y')} — {end.strftime('%d %B %Y')}"
    ws1["A2"].font = SUB_FONT
    ws1.row_dimensions[1].height = 22
    ws1.row_dimensions[3].height = 18

    stats = [
        ("Total kunjungan",    summary["total_visits"]),
        ("Pengunjung unik",    summary["unique_visitors"]),
        ("Rata-rata per hari", summary["avg_per_day"]),
        ("Hari tersibuk",      f"{summary['busiest_day']} ({summary['busiest_count']} kunjungan)"),
        ("Jam tersibuk",       f"{summary['peak_hour']:02d}:00 ({summary['peak_hour_count']} kunjungan)"),
        ("Hari aktif",         summary["days_with_data"]),
        ("Anggota tidak aktif", len(get_inactive_members())),
    ]
    _hdr_row(ws1, 4, ["Statistik", "Nilai"])
    for ri, (label, val) in enumerate(stats, 5):
        _data_row(ws1, ri, [label, val], alt=(ri % 2 == 0))
    _autowidth(ws1)

    # ── Sheet 2: Detail kunjungan ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Detail Kunjungan")
    detail = get_all_visits_detail(start, end)
    _hdr_row(ws2, 1, ["No", "Nama", "ID Barcode", "Tanggal", "Waktu Masuk"])
    for ri, row in enumerate(detail, 2):
        _data_row(ws2, ri, [
            ri - 1,
            row["nama"],
            row["barcode_id"],
            row["tanggal"],
            row["waktu_masuk"][11:19] if len(row["waktu_masuk"]) >= 19 else row["waktu_masuk"],
        ], alt=(ri % 2 == 0))
    _autowidth(ws2)

    # ── Sheet 3: Top anggota ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Top Anggota")
    top = get_top_visitors(start, end, limit=50)
    _hdr_row(ws3, 1, ["Rank", "Nama", "ID Barcode", "Jumlah Kunjungan", "% dari Total"])
    for ri, row in enumerate(top, 2):
        _data_row(ws3, ri, [
            row["rank"], row["nama"], row["barcode_id"],
            row["count"], f"{row['pct']}%"
        ], alt=(ri % 2 == 0))
    _autowidth(ws3)

    # ── Sheet 4: Tren harian ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("Tren Harian")
    trend = get_daily_trend(start, end)
    _hdr_row(ws4, 1, ["Tanggal", "Hari", "Jumlah Kunjungan"])
    for ri, row in enumerate(trend, 2):
        _data_row(ws4, ri, [row["tanggal"], row["hari"], row["count"]], alt=(ri % 2 == 0))
    _autowidth(ws4)

    # ── Sheet 5: Distribusi per jam ───────────────────────────────────────────
    ws5 = wb.create_sheet("Per Jam")
    hourly = get_hourly_distribution(start, end)
    _hdr_row(ws5, 1, ["Jam", "Jumlah Kunjungan"])
    for ri, row in enumerate(hourly, 2):
        _data_row(ws5, ri, [row["label"], row["count"]], alt=(ri % 2 == 0))
    _autowidth(ws5)

    # ── Sheet 6: Tidak aktif ──────────────────────────────────────────────────
    ws6 = wb.create_sheet("Tidak Aktif")
    inactive = get_inactive_members()
    # Ambil semua kolom yang ada
    all_keys = list(inactive[0].keys()) if inactive else [EXCEL_COL_BARCODE, EXCEL_COL_NAME]
    _hdr_row(ws6, 1, ["No"] + all_keys)
    for ri, row in enumerate(inactive, 2):
        _data_row(ws6, ri, [ri - 1] + [str(row.get(k, "") or "") for k in all_keys],
                  alt=(ri % 2 == 0))
    _autowidth(ws6)

    wb.save(out_path)
    logger.info("Export Excel: %d baris → %s", len(detail), out_path)
    return len(detail)


# ══════════════════════════════════════════════════════════════════════════════
# Export PDF
# ══════════════════════════════════════════════════════════════════════════════

def export_to_pdf(start: date, end: date, out_path: str) -> bool:
    """
    Export laporan ringkasan ke PDF menggunakan reportlab.

    Returns True jika berhasil.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        summary = get_summary(start, end)
        top     = get_top_visitors(start, end, limit=10)
        trend   = get_daily_trend(start, end)
        hourly  = get_hourly_distribution(start, end)
        inactive = get_inactive_members()

        P1 = colors.HexColor("#4A5E3A")
        P2 = colors.HexColor("#6B7F58")
        P4 = colors.HexColor("#D4C87A")
        BG = colors.HexColor("#E8E3D8")
        WH = colors.white

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("title",
            fontSize=18, fontName="Helvetica-Bold",
            textColor=P1, spaceAfter=4, alignment=TA_CENTER)
        sub_style = ParagraphStyle("sub",
            fontSize=11, fontName="Helvetica",
            textColor=P2, spaceAfter=16, alignment=TA_CENTER)
        section_style = ParagraphStyle("section",
            fontSize=12, fontName="Helvetica-Bold",
            textColor=P1, spaceBefore=14, spaceAfter=6)

        def _tbl_style(hdr_rows=1):
            return TableStyle([
                ("BACKGROUND",  (0,0), (-1, hdr_rows-1), P1),
                ("TEXTCOLOR",   (0,0), (-1, hdr_rows-1), WH),
                ("FONTNAME",    (0,0), (-1, hdr_rows-1), "Helvetica-Bold"),
                ("FONTSIZE",    (0,0), (-1,-1), 9),
                ("ROWBACKGROUNDS", (0, hdr_rows), (-1,-1), [WH, BG]),
                ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#C5B99A")),
                ("ALIGN",       (0,0), (-1,-1), "CENTER"),
                ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
                ("TOPPADDING",  (0,0), (-1,-1), 4),
                ("BOTTOMPADDING",(0,0),(-1,-1), 4),
                ("LEFTPADDING", (0,0), (-1,-1), 6),
            ])

        doc   = SimpleDocTemplate(out_path, pagesize=A4,
                    leftMargin=2*cm, rightMargin=2*cm,
                    topMargin=2*cm, bottomMargin=2*cm)
        story = []

        # Judul
        story.append(Paragraph("Laporan Perpustakaan", title_style))
        story.append(Paragraph(
            f"Periode: {start.strftime('%d %B %Y')} — {end.strftime('%d %B %Y')}",
            sub_style))
        story.append(HRFlowable(width="100%", thickness=1, color=P1))
        story.append(Spacer(1, 12))

        # Ringkasan statistik
        story.append(Paragraph("Ringkasan", section_style))
        stat_data = [
            ["Statistik", "Nilai"],
            ["Total kunjungan",      str(summary["total_visits"])],
            ["Pengunjung unik",      str(summary["unique_visitors"])],
            ["Rata-rata per hari",   str(summary["avg_per_day"])],
            ["Hari tersibuk",        f"{summary['busiest_day']} ({summary['busiest_count']} kunjungan)"],
            ["Jam tersibuk",         f"{summary['peak_hour']:02d}:00 ({summary['peak_hour_count']} kunjungan)"],
            ["Hari aktif",           str(summary["days_with_data"])],
            ["Anggota tidak aktif",  str(len(inactive))],
        ]
        t = Table(stat_data, colWidths=[9*cm, 8*cm])
        t.setStyle(_tbl_style())
        story.append(t)
        story.append(Spacer(1, 10))

        # Top 10 anggota
        story.append(Paragraph("Top 10 Anggota Terbanyak", section_style))
        top_data = [["Rank", "Nama", "ID Barcode", "Kunjungan", "%"]]
        for r in top:
            top_data.append([r["rank"], r["nama"], r["barcode_id"], r["count"], f"{r['pct']}%"])
        t2 = Table(top_data, colWidths=[1.5*cm, 7*cm, 4*cm, 2.5*cm, 2*cm])
        t2.setStyle(_tbl_style())
        story.append(t2)
        story.append(Spacer(1, 10))

        # Distribusi per jam (ringkas: hanya jam 07–18)
        story.append(Paragraph("Distribusi Kunjungan per Jam (07:00–18:00)", section_style))
        hour_data = [["Jam", "Jumlah Kunjungan"]] + [
            [h["label"], h["count"]] for h in hourly if 7 <= h["jam"] <= 18
        ]
        t3 = Table(hour_data, colWidths=[6*cm, 6*cm])
        t3.setStyle(_tbl_style())
        story.append(t3)
        story.append(Spacer(1, 10))

        # Anggota tidak aktif (ringkas, max 20)
        if inactive:
            story.append(Paragraph(f"Anggota Tidak Aktif ({len(inactive)} orang)", section_style))
            keys = list(inactive[0].keys())
            ia_data = [["No"] + keys]
            for i, m in enumerate(inactive[:20], 1):
                ia_data.append([i] + [str(m.get(k, "") or "") for k in keys])
            col_w = 17 * cm / (len(keys) + 1)
            t4 = Table(ia_data, colWidths=[col_w] * (len(keys) + 1))
            t4.setStyle(_tbl_style())
            story.append(t4)
            if len(inactive) > 20:
                story.append(Paragraph(
                    f"... dan {len(inactive)-20} anggota lainnya (lihat export Excel untuk data lengkap).",
                    ParagraphStyle("note", fontSize=8, textColor=P2)))

        # Footer
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width="100%", thickness=0.5, color=P2))
        story.append(Paragraph(
            f"Dicetak: {datetime.now().strftime('%d %B %Y %H:%M')}  •  PerpusReworked",
            ParagraphStyle("footer", fontSize=8, textColor=P2, alignment=TA_CENTER, spaceBefore=6)))

        doc.build(story)
        logger.info("Export PDF → %s", out_path)
        return True

    except ImportError:
        logger.error("reportlab tidak terinstall. Jalankan: pip install reportlab")
        return False
    except Exception as exc:
        logger.error("export_to_pdf error: %s", exc)
        return False