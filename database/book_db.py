"""
database/book_db.py — Handler SQLite untuk data buku perpustakaan
"""

import sqlite3
import logging
from typing import Optional

from config import DATABASE_PATH

logger = logging.getLogger(__name__)


def _get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DATABASE_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_book_db() -> None:
    """
    Buat tabel buku jika belum ada.
    Aman dipanggil berulang kali (idempotent).
    Dipanggil dari init_db() di sqlite_db.py.
    """
    ddl = """
    CREATE TABLE IF NOT EXISTS buku (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        kode_buku       TEXT    NOT NULL UNIQUE,
        judul           TEXT    NOT NULL,
        pengarang       TEXT    NOT NULL DEFAULT '',
        isbn            TEXT    NOT NULL DEFAULT '',
        kategori        TEXT    NOT NULL DEFAULT '',
        penerbit        TEXT    NOT NULL DEFAULT '',
        tahun_terbit    TEXT    NOT NULL DEFAULT '',
        jumlah_eksemplar INTEGER NOT NULL DEFAULT 1,
        jumlah_tersedia  INTEGER NOT NULL DEFAULT 1,
        lokasi_rak      TEXT    NOT NULL DEFAULT '',
        keterangan      TEXT    NOT NULL DEFAULT '',
        tanggal_masuk   DATE    NOT NULL DEFAULT (date('now'))
    );

    CREATE INDEX IF NOT EXISTS idx_kode_buku ON buku (kode_buku);
    CREATE INDEX IF NOT EXISTS idx_judul     ON buku (judul);
    CREATE INDEX IF NOT EXISTS idx_kategori  ON buku (kategori);
    """
    try:
        with _get_connection() as conn:
            conn.executescript(ddl)
        logger.info("Tabel buku diinisialisasi.")
    except sqlite3.Error as exc:
        logger.error("Gagal inisialisasi tabel buku: %s", exc)
        raise


# ── CRUD ──────────────────────────────────────────────────────────────────────

def add_book(data: dict) -> tuple[bool, str]:
    """
    Tambah satu buku baru.

    Args:
        data: dict dengan key sesuai kolom tabel.
              Wajib: kode_buku, judul
              Opsional: pengarang, isbn, kategori, penerbit,
                        tahun_terbit, jumlah_eksemplar, lokasi_rak, keterangan

    Returns:
        (True, kode_buku)  — berhasil
        (False, pesan_error) — gagal
    """
    required = ("kode_buku", "judul")
    for key in required:
        if not data.get(key, "").strip():
            return False, f"Field '{key}' wajib diisi."

    kode = data["kode_buku"].strip().upper()
    jml  = max(1, int(data.get("jumlah_eksemplar", 1)))

    sql = """
    INSERT INTO buku
        (kode_buku, judul, pengarang, isbn, kategori, penerbit,
         tahun_terbit, jumlah_eksemplar, jumlah_tersedia, lokasi_rak, keterangan)
    VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """
    params = (
        kode,
        data["judul"].strip(),
        data.get("pengarang", "").strip(),
        data.get("isbn", "").strip(),
        data.get("kategori", "").strip(),
        data.get("penerbit", "").strip(),
        data.get("tahun_terbit", "").strip(),
        jml,
        jml,   # tersedia = total saat pertama masuk
        data.get("lokasi_rak", "").strip(),
        data.get("keterangan", "").strip(),
    )
    try:
        with _get_connection() as conn:
            conn.execute(sql, params)
        logger.info("Buku ditambahkan: %s — %s", kode, data["judul"])
        return True, kode
    except sqlite3.IntegrityError:
        return False, f"Kode buku '{kode}' sudah terdaftar."
    except sqlite3.Error as exc:
        logger.error("Gagal tambah buku: %s", exc)
        return False, str(exc)


def update_book(kode_buku: str, data: dict) -> tuple[bool, str]:
    """
    Update data buku berdasarkan kode_buku.
    Tidak mengubah jumlah_tersedia (dikelola oleh modul peminjaman).
    """
    kode = kode_buku.strip().upper()
    jml  = max(1, int(data.get("jumlah_eksemplar", 1)))

    sql = """
    UPDATE buku SET
        judul            = ?,
        pengarang        = ?,
        isbn             = ?,
        kategori         = ?,
        penerbit         = ?,
        tahun_terbit     = ?,
        jumlah_eksemplar = ?,
        lokasi_rak       = ?,
        keterangan       = ?
    WHERE kode_buku = ?
    """
    params = (
        data.get("judul", "").strip(),
        data.get("pengarang", "").strip(),
        data.get("isbn", "").strip(),
        data.get("kategori", "").strip(),
        data.get("penerbit", "").strip(),
        data.get("tahun_terbit", "").strip(),
        jml,
        data.get("lokasi_rak", "").strip(),
        data.get("keterangan", "").strip(),
        kode,
    )
    try:
        with _get_connection() as conn:
            cur = conn.execute(sql, params)
            if cur.rowcount == 0:
                return False, f"Buku dengan kode '{kode}' tidak ditemukan."
        logger.info("Buku diperbarui: %s", kode)
        return True, kode
    except sqlite3.Error as exc:
        logger.error("Gagal update buku: %s", exc)
        return False, str(exc)


def delete_book(kode_buku: str) -> tuple[bool, str]:
    """Hapus buku berdasarkan kode_buku."""
    kode = kode_buku.strip().upper()
    try:
        with _get_connection() as conn:
            cur = conn.execute("DELETE FROM buku WHERE kode_buku = ?", (kode,))
            if cur.rowcount == 0:
                return False, f"Buku dengan kode '{kode}' tidak ditemukan."
        logger.info("Buku dihapus: %s", kode)
        return True, kode
    except sqlite3.Error as exc:
        logger.error("Gagal hapus buku: %s", exc)
        return False, str(exc)

def delete_books_bulk(kodes: list[str]) -> tuple[bool, str]:
    """
    Menghapus banyak buku sekaligus berdasarkan daftar kode buku.
    Return: (success_boolean, error_message)
    """
    if not kodes:
        return True, "Tidak ada buku yang dipilih."

    # Buat placeholder '? , ? , ?' sesuai dengan jumlah data di list 'kodes'
    placeholders = ",".join("?" for _ in kodes)
    sql = f"DELETE FROM buku WHERE kode_buku IN ({placeholders})"

    try:
        with _get_connection() as conn:
            conn.execute(sql, kodes)
        logger.info("Berhasil menghapus %d buku massal.", len(kodes))
        return True, ""
    except sqlite3.Error as exc:
        logger.error("Gagal menghapus buku massal: %s", exc)
        return False, f"Error database: {exc}"

# ── Query ─────────────────────────────────────────────────────────────────────

def get_all_books() -> list[dict]:
    """Ambil semua buku, urut abjad judul."""
    sql = """
    SELECT id, kode_buku, judul, pengarang, isbn, kategori, penerbit,
           tahun_terbit, jumlah_eksemplar, jumlah_tersedia, lokasi_rak,
           keterangan, tanggal_masuk
    FROM buku
    ORDER BY judul COLLATE NOCASE
    """
    try:
        with _get_connection() as conn:
            rows = conn.execute(sql).fetchall()
        return [dict(r) for r in rows]
    except sqlite3.Error as exc:
        logger.error("Gagal ambil semua buku: %s", exc)
        return []


def search_books(query: str) -> list[dict]:
    """
    Cari buku berdasarkan judul, pengarang, kode_buku, atau ISBN.
    Partial match, case-insensitive.
    """
    q = f"%{query.strip()}%"
    sql = """
    SELECT id, kode_buku, judul, pengarang, isbn, kategori, penerbit,
           tahun_terbit, jumlah_eksemplar, jumlah_tersedia, lokasi_rak,
           keterangan, tanggal_masuk
    FROM buku
    WHERE judul      LIKE ? COLLATE NOCASE
       OR pengarang  LIKE ? COLLATE NOCASE
       OR kode_buku  LIKE ? COLLATE NOCASE
       OR isbn       LIKE ? COLLATE NOCASE
    ORDER BY judul COLLATE NOCASE
    """
    try:
        with _get_connection() as conn:
            rows = conn.execute(sql, (q, q, q, q)).fetchall()
        return [dict(r) for r in rows]
    except sqlite3.Error as exc:
        logger.error("Gagal cari buku: %s", exc)
        return []


def get_book_by_kode(kode_buku: str) -> Optional[dict]:
    """Cari satu buku berdasarkan kode (exact match)."""
    kode = kode_buku.strip().upper()
    sql = "SELECT * FROM buku WHERE kode_buku = ?"
    try:
        with _get_connection() as conn:
            row = conn.execute(sql, (kode,)).fetchone()
        return dict(row) if row else None
    except sqlite3.Error as exc:
        logger.error("Gagal cari buku by kode: %s", exc)
        return None


def get_categories() -> list[str]:
    """Ambil daftar kategori unik untuk ComboBox filter."""
    sql = "SELECT DISTINCT kategori FROM buku WHERE kategori != '' ORDER BY kategori"
    try:
        with _get_connection() as conn:
            rows = conn.execute(sql).fetchall()
        return [r[0] for r in rows]
    except sqlite3.Error as exc:
        logger.error("Gagal ambil kategori: %s", exc)
        return []


def get_book_stats() -> dict:
    """Stat ringkas untuk dashboard: total judul, total eksemplar, tersedia."""
    sql = """
    SELECT
        COUNT(*)                  AS total_judul,
        SUM(jumlah_eksemplar)     AS total_eksemplar,
        SUM(jumlah_tersedia)      AS total_tersedia
    FROM buku
    """
    try:
        with _get_connection() as conn:
            row = conn.execute(sql).fetchone()
        return {
            "total_judul"     : row["total_judul"]     or 0,
            "total_eksemplar" : row["total_eksemplar"] or 0,
            "total_tersedia"  : row["total_tersedia"]  or 0,
        }
    except sqlite3.Error as exc:
        logger.error("Gagal ambil statistik buku: %s", exc)
        return {"total_judul": 0, "total_eksemplar": 0, "total_tersedia": 0}


# ── Import massal dari Excel ──────────────────────────────────────────────────

def import_books_from_excel(path: str) -> tuple[int, int, list[str]]:
    """
    Import buku massal dari file Excel.

    Kolom wajib   : Kode Buku, Judul
    Kolom opsional: Pengarang, ISBN, Kategori, Penerbit,
                    Tahun Terbit, Jumlah Eksemplar, Lokasi Rak, Keterangan

    Returns:
        (berhasil, gagal, list_pesan_error)
    """
    import openpyxl
    from pathlib import Path

    path = Path(path)
    if not path.exists():
        return 0, 0, [f"File tidak ditemukan: {path}"]

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        headers = [
            str(c.value).strip() if c.value else ""
            for c in next(ws.iter_rows(max_row=1))
        ]
    except Exception as exc:
        return 0, 0, [f"Gagal baca file: {exc}"]

    required_cols = ("Kode Buku", "Judul")
    for col in required_cols:
        if col not in headers:
            wb.close()
            return 0, 0, [
                f"Kolom '{col}' tidak ditemukan. Kolom tersedia: {headers}"
            ]

    ok_count  = 0
    err_count = 0
    errors    = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rec = dict(zip(headers, row))

        kode  = str(rec.get("Kode Buku", "") or "").strip().upper()
        judul = str(rec.get("Judul", "") or "").strip()

        if not kode or not judul:
            continue  # skip baris kosong

        data = {
            "kode_buku"        : kode,
            "judul"            : judul,
            "pengarang"        : str(rec.get("Pengarang",       "") or "").strip(),
            "isbn"             : str(rec.get("ISBN",            "") or "").strip(),
            "kategori"         : str(rec.get("Kategori",        "") or "").strip(),
            "penerbit"         : str(rec.get("Penerbit",        "") or "").strip(),
            "tahun_terbit"     : str(rec.get("Tahun Terbit",    "") or "").strip(),
            "jumlah_eksemplar" : int(rec.get("Jumlah Eksemplar", 1) or 1),
            "lokasi_rak"       : str(rec.get("Lokasi Rak",      "") or "").strip(),
            "keterangan"       : str(rec.get("Keterangan",      "") or "").strip(),
        }

        success, msg = add_book(data)
        if success:
            ok_count += 1
        else:
            err_count += 1
            errors.append(f"Baris {row_num}: {msg}")

    wb.close()
    logger.info("Import Excel selesai: %d berhasil, %d gagal", ok_count, err_count)
    return ok_count, err_count, errors


def generate_kode_buku(prefix: str = "BK") -> str:
    """
    Auto-generate kode buku berikutnya dalam format PREFIX-YYYY-NNNN.
    Contoh: BK-2026-0001, BK-2026-0002, dst.
    """
    from datetime import date
    year = date.today().year
    sql = """
    SELECT kode_buku FROM buku
    WHERE kode_buku LIKE ?
    ORDER BY kode_buku DESC
    LIMIT 1
    """
    pattern = f"{prefix}-{year}-%"
    try:
        with _get_connection() as conn:
            row = conn.execute(sql, (pattern,)).fetchone()

        if row:
            last = row[0]          # contoh: "BK-2026-0042"
            last_num = int(last.split("-")[-1])
            next_num = last_num + 1
        else:
            next_num = 1

        return f"{prefix}-{year}-{next_num:04d}"
    except Exception:
        return f"{prefix}-{year}-0001"
