"""
config.py — Konfigurasi global PerpusReworked
Semua path dihitung relatif terhadap lokasi aplikasi berjalan
sehingga tetap portable setelah di-build (Nuitka --standalone/--onefile).
"""

from pathlib import Path


def _resolve_base_dir() -> Path:
    """Tentukan folder dasar untuk data persisten (database, anggota.xlsx, backup).

    - Dijalankan sebagai skrip Python biasa (`python main.py`):
      pakai folder tempat config.py berada, seperti biasa.
    - Dikompilasi dengan Nuitka (--standalone atau --onefile):
      pakai `__compiled__.containing_dir`, yaitu folder tempat .exe
      SEBENARNYA berada. Ini WAJIB untuk mode --onefile, karena
      `__file__` di sana menunjuk ke folder temp ekstraksi sementara
      (%TEMP%\\onefile_...) yang dibuat ulang/dihapus tiap kali .exe
      dijalankan — kalau dipakai, database & data anggota akan
      "hilang" setiap aplikasi ditutup lalu dibuka lagi.
    """
    try:
        # Tersedia otomatis saat dikompilasi dengan Nuitka
        return Path(__compiled__.containing_dir).resolve()  # type: ignore[name-defined]
    except NameError:
        return Path(__file__).parent.resolve()


# ── Root direktori proyek ─────────────────────────────────────────────────────
BASE_DIR: Path = _resolve_base_dir()

# ── Path data & database ──────────────────────────────────────────────────────
DATA_DIR: Path       = BASE_DIR / "data"
DATABASE_DIR: Path   = BASE_DIR / "database"
ASSETS_DIR: Path     = BASE_DIR / "assets"

MEMBER_EXCEL_PATH: Path  = DATA_DIR / "anggota.xlsx"
DATABASE_PATH: Path      = DATABASE_DIR / "perpus.db"
BACKUP_DIR: Path         = DATA_DIR / "backup"

# ── Kolom wajib di anggota.xlsx ───────────────────────────────────────────────
EXCEL_COL_BARCODE: str = "ID Barcode"
EXCEL_COL_NAME: str    = "Nama"

# ── Kamera ───────────────────────────────────────────────────────────────────
CAMERA_SCAN_RANGE: int     = 10       # Indeks kamera yang discan (0–9)
CAMERA_FRAME_WIDTH: int    = 640
CAMERA_FRAME_HEIGHT: int   = 480
CAMERA_FPS: int            = 30

# ── Barcode scanner ───────────────────────────────────────────────────────────
BARCODE_DEBOUNCE_SECONDS: float = 2.0   # Jeda minimum antar scan barcode sama
# Format barcode/QR yang dicari scanner. Pisahkan dengan koma untuk beberapa
# format sekaligus, mis. "CODE128,QRCODE".
BARCODE_TYPE_FILTER: str        = "QRCODE"

# ── GUI ───────────────────────────────────────────────────────────────────────
APP_NAME: str    = "PerpusReworked"
APP_VERSION: str = "1.0.0"
WINDOW_MIN_WIDTH: int  = 1024
WINDOW_MIN_HEIGHT: int = 640

# Timer interval update frame kamera (milidetik)
CAMERA_TIMER_INTERVAL_MS: int = 33     # ~30 fps

# ── Pastikan direktori penting ada saat import ────────────────────────────────
for _dir in (DATA_DIR, DATABASE_DIR, ASSETS_DIR, BACKUP_DIR):
    _dir.mkdir(parents=True, exist_ok=True)

# Log directory (dibuat saat runtime di main.py)
LOG_DIR: Path = DATA_DIR / "logs"


# ── Pastikan anggota.xlsx ada dengan header yang benar ───────────────────────
def _ensure_member_excel():
    """Buat anggota.xlsx dengan header default jika belum ada."""
    if MEMBER_EXCEL_PATH.exists():
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Anggota"

        # Header dengan styling
        headers = [EXCEL_COL_BARCODE, EXCEL_COL_NAME]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill("solid", fgColor="4A5E3A")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(str(MEMBER_EXCEL_PATH))
    except ImportError:
        # openpyxl belum terinstall, abaikan untuk sekarang
        pass


_ensure_member_excel()